from __future__ import annotations

import pandas as pd
from openpyxl.utils.cell import get_column_letter
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QComboBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..plan_config import SheetRange

_PREVIEW_MAX_COL_WIDTH = 420


class SheetExcelPreview(QWidget):
    """原始 Excel 网格预览，支持普通区域和规划表双区域取点。"""

    range_changed = Signal(str, str)  # start_cell, end_cell
    region_changed = Signal(str, str, str)  # region_key, start_cell, end_cell

    _MAX_ROWS = 500
    _MAX_COLS = 120

    _REGIONS = {
        "table": ("普通执行区域", QColor(198, 239, 206)),
        "mechanism": ("机制信息区域", QColor(221, 235, 247)),
        "product": ("商品勾选区域", QColor(255, 242, 204)),
    }

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._dataframe: pd.DataFrame | None = None
        self._ranges: dict[str, SheetRange] = {
            "table": SheetRange("A1", ""),
            "mechanism": SheetRange("D2", ""),
            "product": SheetRange("C10", ""),
        }
        self._active_region = "table"
        self._pick_edge = "start"
        self._show_plan_regions = False

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        toolbar = QHBoxLayout()
        self.info_label = QLabel("请在左侧 Sheet 列表中选择一行以预览内容。")
        self.region_combo = QComboBox()
        for key, (label, _color) in self._REGIONS.items():
            self.region_combo.addItem(label, key)
        self.region_combo.currentIndexChanged.connect(self._on_region_combo_changed)

        self.set_start_btn = QPushButton("设左上角")
        self.set_end_btn = QPushButton("设右下角")
        self.set_start_btn.setCheckable(True)
        self.set_end_btn.setCheckable(True)
        self.set_start_btn.setChecked(True)
        self.set_start_btn.clicked.connect(lambda: self.set_region_pick_mode(self._active_region, "start"))
        self.set_end_btn.clicked.connect(lambda: self.set_region_pick_mode(self._active_region, "end"))
        self.clear_btn = QPushButton("清空当前区域")
        self.clear_btn.clicked.connect(self._clear_active_region)

        toolbar.addWidget(self.info_label, stretch=1)
        toolbar.addWidget(QLabel("取点："))
        toolbar.addWidget(self.region_combo)
        toolbar.addWidget(self.set_start_btn)
        toolbar.addWidget(self.set_end_btn)
        toolbar.addWidget(self.clear_btn)
        layout.addLayout(toolbar)

        labels = QHBoxLayout()
        self.table_label = QLabel()
        self.mechanism_label = QLabel()
        self.product_label = QLabel()
        labels.addWidget(self.table_label)
        labels.addWidget(self.mechanism_label)
        labels.addWidget(self.product_label)
        labels.addStretch()
        layout.addLayout(labels)

        self.grid = QTableWidget(0, 0)
        self.grid.setAlternatingRowColors(True)
        configure_table_full_text(self.grid)
        self.grid.cellClicked.connect(self._on_cell_clicked)
        layout.addWidget(self.grid, stretch=1)

        hint = QLabel("点击「设左上角」或「设右下角」后，直接在表格中点单元格；规划表建议先框机制信息区，再框商品勾选区。")
        hint.setStyleSheet("color: #666;")
        layout.addWidget(hint)

        self._update_range_labels()

    def load_sheet(
        self,
        file_name: str,
        sheet_name: str,
        dataframe: pd.DataFrame,
        sheet_range: SheetRange | None = None,
        plan_ranges: dict[str, SheetRange] | None = None,
        active_region: str = "table",
        show_plan_regions: bool = False,
    ) -> None:
        self._dataframe = dataframe
        self._show_plan_regions = show_plan_regions
        self._ranges["table"] = sheet_range or SheetRange("A1", "")
        if plan_ranges:
            for key in ("mechanism", "product"):
                if key in plan_ranges:
                    self._ranges[key] = plan_ranges[key]

        self.set_region_pick_mode(active_region, self._pick_edge, rerender=False)
        self.info_label.setText(f"{file_name} / {sheet_name}")
        self._render_grid()
        self._update_range_labels()

    def clear(self) -> None:
        self._dataframe = None
        self.grid.setRowCount(0)
        self.grid.setColumnCount(0)
        self.info_label.setText("请在左侧 Sheet 列表中选择一行以预览内容。")

    def current_range(self) -> SheetRange:
        return self._ranges["table"]

    def current_region_range(self, region: str) -> SheetRange:
        return self._ranges.get(region, SheetRange())

    def set_range(self, start_cell: str, end_cell: str) -> None:
        self.set_region_range("table", SheetRange(start_cell or "A1", end_cell or ""))

    def set_region_range(self, region: str, sheet_range: SheetRange) -> None:
        if region not in self._REGIONS:
            return
        self._ranges[region] = sheet_range
        self._render_grid()
        self._update_range_labels()
        self._emit_region(region)

    def set_region_pick_mode(self, region: str, edge: str = "start", rerender: bool = True) -> None:
        if region not in self._REGIONS:
            region = "table"
        self._active_region = region
        self._pick_edge = "end" if edge == "end" else "start"
        self.set_start_btn.setChecked(self._pick_edge == "start")
        self.set_end_btn.setChecked(self._pick_edge == "end")
        combo_index = self.region_combo.findData(region)
        if combo_index >= 0 and self.region_combo.currentIndex() != combo_index:
            self.region_combo.blockSignals(True)
            self.region_combo.setCurrentIndex(combo_index)
            self.region_combo.blockSignals(False)
        if rerender:
            self._render_grid()
            self._update_range_labels()

    def set_plan_region_visibility(self, visible: bool) -> None:
        self._show_plan_regions = visible
        self._render_grid()
        self._update_range_labels()

    def _on_region_combo_changed(self) -> None:
        region = self.region_combo.currentData()
        self.set_region_pick_mode(str(region), self._pick_edge)

    def _clear_active_region(self) -> None:
        defaults = {
            "table": SheetRange("A1", ""),
            "mechanism": SheetRange("D2", ""),
            "product": SheetRange("C10", ""),
        }
        self.set_region_range(self._active_region, defaults[self._active_region])

    def _on_cell_clicked(self, row: int, column: int) -> None:
        if self._dataframe is None:
            return
        cell = f"{get_column_letter(column + 1)}{row + 1}"
        current = self._ranges[self._active_region]
        if self._pick_edge == "start":
            self._ranges[self._active_region] = SheetRange(cell, current.end_cell)
        else:
            self._ranges[self._active_region] = SheetRange(current.start_cell or "A1", cell)
        self._render_grid()
        self._update_range_labels()
        self._emit_region(self._active_region)

    def _emit_region(self, region: str) -> None:
        sheet_range = self._ranges[region]
        self.region_changed.emit(region, sheet_range.start_cell, sheet_range.end_cell)
        if region == "table":
            self.range_changed.emit(sheet_range.start_cell, sheet_range.end_cell)

    def _render_grid(self) -> None:
        if self._dataframe is None or self._dataframe.empty:
            self.grid.setRowCount(0)
            self.grid.setColumnCount(0)
            return

        row_count = min(len(self._dataframe), self._MAX_ROWS)
        col_count = min(len(self._dataframe.columns), self._MAX_COLS)
        self.grid.setRowCount(row_count)
        self.grid.setColumnCount(col_count)

        self.grid.setVerticalHeaderLabels([str(i + 1) for i in range(row_count)])
        self.grid.setHorizontalHeaderLabels([get_column_letter(i + 1) for i in range(col_count)])

        visible_regions = ["table"]
        if self._show_plan_regions:
            visible_regions = ["mechanism", "product"]

        for row_index in range(row_count):
            for column_index in range(col_count):
                value = self._dataframe.iloc[row_index, column_index]
                text = "" if pd.isna(value) else str(value).strip()
                item = QTableWidgetItem(text)
                if text:
                    item.setToolTip(text)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

                excel_row = row_index + 1
                excel_col = column_index + 1
                for region in visible_regions:
                    if _cell_in_range(excel_row, excel_col, self._ranges[region]):
                        _label, color = self._REGIONS[region]
                        item.setBackground(color)
                        item.setToolTip(_label if not text else f"{_label}\n{text}")
                if _cell_in_range(excel_row, excel_col, self._ranges[self._active_region]):
                    item.setForeground(QColor(0, 0, 0))

                self.grid.setItem(row_index, column_index, item)

        autosize_table_columns(self.grid, max_width=_PREVIEW_MAX_COL_WIDTH)

    def _update_range_labels(self) -> None:
        self.table_label.setText(f"普通：{_format_range(self._ranges['table'])}")
        self.mechanism_label.setText(f"机制：{_format_range(self._ranges['mechanism'])}")
        self.product_label.setText(f"商品：{_format_range(self._ranges['product'])}")
        self.mechanism_label.setVisible(self._show_plan_regions)
        self.product_label.setVisible(self._show_plan_regions)


def configure_table_full_text(table: QTableWidget) -> None:
    table.setWordWrap(True)
    table.setTextElideMode(Qt.ElideNone)
    table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
    table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
    header = table.horizontalHeader()
    header.setSectionResizeMode(QHeaderView.Interactive)
    header.setStretchLastSection(False)
    header.setMaximumSectionSize(2000)


def autosize_table_columns(table: QTableWidget, max_width: int = 420) -> None:
    table.resizeColumnsToContents()
    for column_index in range(table.columnCount()):
        width = min(table.columnWidth(column_index), max_width)
        table.setColumnWidth(column_index, max(width, 48))
    table.resizeRowsToContents()


def _parse_cell(cell: str) -> tuple[int, int]:
    from openpyxl.utils.cell import coordinate_to_tuple

    row, col = coordinate_to_tuple(cell.upper())
    return row, col


def _cell_in_range(row: int, column: int, sheet_range: SheetRange) -> bool:
    if not sheet_range.start_cell:
        return False
    try:
        start_row, start_col = _parse_cell(sheet_range.start_cell)
        if sheet_range.end_cell:
            end_row, end_col = _parse_cell(sheet_range.end_cell)
        else:
            end_row, end_col = start_row, start_col
    except ValueError:
        return False

    min_row, max_row = sorted((start_row, end_row))
    min_col, max_col = sorted((start_col, end_col))
    return min_row <= row <= max_row and min_col <= column <= max_col


def _format_range(sheet_range: SheetRange) -> str:
    if not sheet_range.start_cell:
        return "未设置"
    if not sheet_range.end_cell:
        return f"{sheet_range.start_cell}:自动"
    return f"{sheet_range.start_cell}:{sheet_range.end_cell}"
