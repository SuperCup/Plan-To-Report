from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from ..app_settings import load_app_settings, save_app_settings, settings_path
from ..engine import ConversionEngine
from ..excel_io import ParsedSheet, parse_workbook_sheets, table_from_range, write_outputs
from ..models import ConversionResult, ConversionTemplate
from ..plan_config import AiSettings, SheetRange
from ..plan_parser import infer_config_from_range
from ..plan_pipeline import PlanConversionInput, output_tables, run_plan_conversion
from ..sheet_inputs import _detect_role_range, load_role_table
from ..template_loader import list_templates, load_template
from .excel_preview import SheetExcelPreview, autosize_table_columns, configure_table_full_text
from .plan_wizard import PlanWizardWidget


class MainWindow(QMainWindow):
    def __init__(self, project_root: Path) -> None:
        super().__init__()
        self.project_root = project_root
        self.template_dir = project_root / "templates"
        self.output_dir = project_root / "output"
        self.current_template: ConversionTemplate | None = None
        self.current_result: ConversionResult | None = None
        self.parsed_sheets: list[ParsedSheet] = []
        self.sheet_ranges: dict[int, SheetRange] = {}
        self.plan_ranges: dict[int, dict[str, SheetRange]] = {}
        self.sheet_roles: dict[int, str] = {}
        self._last_upload_dir = project_root
        self._preview_row_index: int | None = None
        self.app_settings = load_app_settings(project_root)

        self.setWindowTitle("规划转提报工具 - 批量Sheet版")
        self.resize(1680, 960)

        splitter = QSplitter(Qt.Horizontal)
        self.setCentralWidget(splitter)

        splitter.addWidget(self._build_left_panel())

        self.excel_preview = SheetExcelPreview()
        self.excel_preview.region_changed.connect(self._on_preview_region_changed)
        preview_group = QGroupBox("Excel 内容预览")
        preview_layout = QVBoxLayout(preview_group)
        preview_layout.addWidget(self.excel_preview)
        splitter.addWidget(preview_group)

        splitter.addWidget(self._build_right_panel())
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 5)
        splitter.setStretchFactor(2, 2)
        splitter.setSizes([360, 900, 420])

        self.reload_templates()

    def _build_left_panel(self) -> QWidget:
        panel = QWidget()
        panel.setMinimumWidth(320)
        panel.setMaximumWidth(460)
        layout = QVBoxLayout(panel)

        layout.addWidget(self._build_template_bar())
        layout.addWidget(self._build_upload_panel(), stretch=1)

        action_layout = QHBoxLayout()
        self.run_button = QPushButton("执行转换")
        self.run_button.clicked.connect(self.run_conversion)
        self.export_button = QPushButton("导出结果")
        self.export_button.clicked.connect(self.export_outputs)
        self.export_button.setEnabled(False)
        action_layout.addWidget(self.run_button)
        action_layout.addWidget(self.export_button)
        layout.addLayout(action_layout)

        self.status = QLabel("请选择转换模板。")
        self.status.setWordWrap(True)
        self.status.setStyleSheet("color: #333;")
        layout.addWidget(self.status)
        return panel

    def _build_right_panel(self) -> QWidget:
        panel = QWidget()
        panel.setMinimumWidth(360)
        layout = QVBoxLayout(panel)

        layout.addWidget(self._build_execution_panel(), stretch=3)

        self.result_tabs = QTabWidget()
        self.result_tabs.setVisible(False)
        layout.addWidget(self.result_tabs, stretch=2)
        return panel

    def _build_template_bar(self) -> QWidget:
        group = QGroupBox("模板")
        layout = QVBoxLayout(group)
        self.template_combo = QComboBox()
        self.template_combo.currentIndexChanged.connect(self.load_selected_template)
        reload_button = QPushButton("刷新模板")
        reload_button.clicked.connect(self.reload_templates)
        layout.addWidget(self.template_combo)
        layout.addWidget(reload_button)
        return group

    def _build_upload_panel(self) -> QWidget:
        group = QGroupBox("文件与 Sheet")
        layout = QVBoxLayout(group)

        button_layout = QHBoxLayout()
        upload_button = QPushButton("上传 Excel")
        upload_button.clicked.connect(self.choose_excels)
        preview_button = QPushButton("预览选中")
        preview_button.clicked.connect(self._preview_selected_sheet)
        button_layout.addWidget(upload_button)
        button_layout.addWidget(preview_button)
        layout.addLayout(button_layout)

        self.upload_status = QLabel("未上传文件。")
        self.upload_status.setWordWrap(True)
        self.upload_status.setStyleSheet("color: #666;")
        layout.addWidget(self.upload_status)

        self.sheet_table = QTableWidget(0, 8)
        self.sheet_table.setHorizontalHeaderLabels(
            ["文件", "Sheet", "行数", "列数", "合并提示", "Sheet 类型", "机制/起始", "商品/结束"]
        )
        self.sheet_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.sheet_table.setSelectionMode(QTableWidget.SingleSelection)
        configure_table_full_text(self.sheet_table)
        self.sheet_table.itemSelectionChanged.connect(self._on_sheet_selection_changed)
        layout.addWidget(self.sheet_table, stretch=1)
        return group

    def _build_execution_panel(self) -> QWidget:
        group = QGroupBox("配置")
        layout = QVBoxLayout(group)

        self.legacy_execution_panel = QWidget()
        legacy_layout = QFormLayout(self.legacy_execution_panel)
        self.direction_combo = QComboBox()
        self.direction_combo.addItem("逐行执行：首行为字段名，后续每行一条记录", "row")
        self.direction_combo.addItem("逐列执行：首列为字段名，后续每列一条记录", "column")
        legacy_layout.addRow("普通 Sheet 执行顺序：", self.direction_combo)
        legacy_hint = QLabel("普通 Sheet 在中央预览区框选执行区域；规划表请在左侧类型中选「规划表」。")
        legacy_hint.setWordWrap(True)
        legacy_hint.setStyleSheet("color: #666;")
        legacy_layout.addRow(legacy_hint)

        self.plan_wizard = PlanWizardWidget(
            project_root=self.project_root,
            on_save_ai=self._persist_ai_settings,
        )
        self.plan_wizard.pick_requested.connect(self._on_plan_pick_requested)
        self.plan_wizard.load_ai_settings(
            self.app_settings.deepseek,
            settings_path(self.project_root),
        )

        layout.addWidget(self.legacy_execution_panel)
        layout.addWidget(self.plan_wizard, stretch=1)
        self.plan_wizard.hide()
        return group

    def reload_templates(self) -> None:
        self.template_combo.blockSignals(True)
        self.template_combo.clear()
        for path in list_templates(self.template_dir):
            self.template_combo.addItem(path.stem, str(path))
        self.template_combo.blockSignals(False)

        if self.template_combo.count() == 0:
            self.status.setText("未找到模板，请在 templates 目录添加 .json 模板。")
            return
        self.template_combo.setCurrentIndex(0)
        self.load_selected_template()

    def load_selected_template(self) -> None:
        path = self.template_combo.currentData()
        if not path:
            return

        try:
            self.current_template = load_template(path)
        except Exception as exc:
            QMessageBox.critical(self, "模板错误", str(exc))
            return

        self._render_sheet_table()
        self.result_tabs.clear()
        self.result_tabs.setVisible(False)
        self.current_result = None
        self.export_button.setEnabled(False)
        self.status.setText(f"已加载模板：{self.current_template.name} v{self.current_template.version}")

    def choose_excels(self) -> None:
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "选择一个或多个 Excel 文件",
            str(self._last_upload_dir if self._last_upload_dir.exists() else self.project_root),
            "Excel Files (*.xlsx *.xlsm)",
        )
        if not paths:
            return

        self._sync_sheet_roles_from_widgets()
        previous_count = len(self.parsed_sheets)
        self._last_upload_dir = Path(paths[0]).resolve().parent
        parsed: list[ParsedSheet] = []
        errors: list[str] = []
        for path in paths:
            try:
                parsed.extend(parse_workbook_sheets(path))
            except Exception as exc:
                errors.append(f"{Path(path).name}: {exc}")

        self.parsed_sheets.extend(parsed)
        self.current_result = None
        self.export_button.setEnabled(False)
        self.result_tabs.clear()
        self.result_tabs.setVisible(False)
        self._render_sheet_table()
        self.upload_status.setText(
            f"本次新增 {len(paths)} 个文件、{len(parsed)} 个 Sheet；当前共 {len(self.parsed_sheets)} 个 Sheet。"
        )
        if errors:
            QMessageBox.warning(self, "部分文件解析失败", "\n".join(errors))
        if parsed:
            self.sheet_table.selectRow(previous_count)
            self._preview_sheet_at_row(previous_count)

    def _render_sheet_table(self) -> None:
        if not hasattr(self, "sheet_table"):
            return

        self._sync_sheet_roles_from_widgets()
        self.sheet_table.setRowCount(len(self.parsed_sheets))
        role_options = self._role_options()
        for row_index, sheet in enumerate(self.parsed_sheets):
            display_range = self._display_ranges_for_row(row_index)
            values = [
                sheet.file_path.name,
                sheet.sheet_name,
                str(len(sheet.data)),
                str(len(sheet.data.columns)),
                "; ".join(sheet.merge_notes),
                "",
                display_range[0],
                display_range[1],
            ]
            for column_index, value in enumerate(values):
                item = QTableWidgetItem(value)
                if value:
                    item.setToolTip(value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.sheet_table.setItem(row_index, column_index, item)

            combo = QComboBox()
            combo.addItems(role_options)
            role = self.sheet_roles.get(row_index, "")
            if role:
                combo.setCurrentText(role)
            combo.currentTextChanged.connect(lambda _text, idx=row_index: self._on_role_changed(idx))
            self.sheet_table.setCellWidget(row_index, 5, combo)

        autosize_table_columns(self.sheet_table, max_width=260)

    def _on_role_changed(self, row_index: int) -> None:
        combo = self.sheet_table.cellWidget(row_index, 5)
        if isinstance(combo, QComboBox):
            role = combo.currentText().strip()
            self.sheet_roles[row_index] = role
        else:
            role = ""
        if role in {"商品清单", "商品匹配逻辑"}:
            current_range = self.sheet_ranges.get(row_index, SheetRange())
            if (current_range.start_cell or "A1").upper() == "A1" and not current_range.end_cell.strip():
                self.sheet_ranges[row_index] = _detect_role_range(self.parsed_sheets[row_index].data, role)
        if role == "规划表":
            self._ensure_plan_ranges(row_index)
            self.sheet_table.selectRow(row_index)
            self._preview_sheet_at_row(row_index)
            self.status.setText("已指定为规划表：请在中央预览区框选机制信息区和商品勾选区。")
        self._update_execution_mode(row_index)
        self._update_range_cells(row_index)

    def _on_sheet_selection_changed(self) -> None:
        rows = self.sheet_table.selectionModel().selectedRows()
        if not rows:
            return
        self._preview_sheet_at_row(rows[0].row())

    def _preview_selected_sheet(self) -> None:
        rows = self.sheet_table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.information(self, "提示", "请先在 Sheet 列表中选择一行。")
            return
        self._preview_sheet_at_row(rows[0].row())

    def _preview_sheet_at_row(self, row_index: int) -> None:
        if row_index < 0 or row_index >= len(self.parsed_sheets):
            return

        self._preview_row_index = row_index
        sheet = self.parsed_sheets[row_index]
        sheet_range = self.sheet_ranges.get(row_index, SheetRange())
        is_plan = self._is_plan_sheet_row(row_index)
        plan_ranges = self._ensure_plan_ranges(row_index) if is_plan else None
        self.excel_preview.load_sheet(
            sheet.file_path.name,
            sheet.sheet_name,
            sheet.data,
            sheet_range,
            plan_ranges,
            active_region="mechanism" if is_plan else "table",
            show_plan_regions=is_plan,
        )
        self._update_execution_mode(row_index)
        if is_plan and plan_ranges:
            self.plan_wizard.refresh_from_sheet(
                sheet.data,
                plan_ranges["mechanism"],
                plan_ranges["product"],
            )

    def _on_preview_region_changed(self, region: str, start_cell: str, end_cell: str) -> None:
        if self._preview_row_index is None:
            return

        row_index = self._preview_row_index
        if region == "table":
            self.sheet_ranges[row_index] = SheetRange(start_cell=start_cell, end_cell=end_cell)
        else:
            ranges = self._ensure_plan_ranges(row_index)
            ranges[region] = SheetRange(start_cell=start_cell, end_cell=end_cell)
            if self._is_plan_sheet_row(row_index):
                sheet = self.parsed_sheets[row_index]
                self.plan_wizard.refresh_from_sheet(
                    sheet.data,
                    ranges["mechanism"],
                    ranges["product"],
                )
        self._update_range_cells(row_index)

    def _on_plan_pick_requested(self, region: str, edge: str) -> None:
        if self._preview_row_index is None or not self._is_plan_sheet_row(self._preview_row_index):
            self.status.setText("请先在左侧选择一个 Sheet，并将 Sheet 类型设为「规划表」。")
            return
        self.excel_preview.set_region_pick_mode(region, edge)
        region_name = "机制信息区" if region == "mechanism" else "商品勾选区"
        edge_name = "左上角" if edge == "start" else "右下角"
        self.status.setText(f"请在中央预览区点击{region_name}的{edge_name}。")

    def _role_options(self) -> list[str]:
        roles = [
            "",
            "规划表",
            "商品清单",
            "商品匹配逻辑",
            "预算分配表",
            "预算匹配逻辑",
            "门店清单",
            "门店匹配逻辑",
        ]
        if self.current_template:
            roles.extend(spec.key for spec in self.current_template.inputs)
        return list(dict.fromkeys(roles))

    def run_conversion(self) -> None:
        if not self.current_template:
            QMessageBox.warning(self, "提示", "请先选择模板。")
            return
        if not self.parsed_sheets:
            QMessageBox.warning(self, "提示", "请先上传 Excel。")
            return

        plan_index = self._find_plan_sheet_index()
        try:
            if plan_index is not None:
                self.current_result = self._run_plan_conversion(plan_index)
            else:
                input_tables = self._collect_typed_tables()
                engine = ConversionEngine(self.current_template)
                self.current_result = engine.run_tables(input_tables)
        except Exception as exc:
            QMessageBox.critical(self, "执行失败", str(exc))
            return

        self._render_result_preview(self.current_result)
        self.export_button.setEnabled(True)
        issue_count = len(self.current_result.issues)
        self.status.setText(f"转换完成，生成 {len(self.current_result.tables)} 个结果表，发现 {issue_count} 条提示/异常。")

    def _persist_ai_settings(self, ai_settings: AiSettings) -> None:
        self.app_settings.deepseek = ai_settings
        path = save_app_settings(self.project_root, self.app_settings)
        self.plan_wizard.load_ai_settings(ai_settings, path)
        self.status.setText(f"已保存 DeepSeek 配置：{path}")

    def _run_plan_conversion(self, plan_index: int) -> ConversionResult:
        ready, message = self.plan_wizard.is_ready()
        if not ready:
            raise ValueError(message)

        self._persist_ai_settings(self.plan_wizard.get_ai_settings())

        sheet = self.parsed_sheets[plan_index]
        config = self.plan_wizard.get_config()
        logic_table = load_role_table(
            self.parsed_sheets,
            self.sheet_table,
            self.sheet_ranges,
            "商品匹配逻辑",
        )
        product_table = load_role_table(
            self.parsed_sheets,
            self.sheet_table,
            self.sheet_ranges,
            "商品清单",
        )
        pipeline_output = run_plan_conversion(
            PlanConversionInput(
                plan_raw=sheet.data,
                config=config,
                logic_table=logic_table,
                product_table=product_table,
                project_root=self.project_root,
            )
        )
        return ConversionResult(
            tables=output_tables(pipeline_output),
            issues=pipeline_output.issues,
        )

    def _find_plan_sheet_index(self) -> int | None:
        if self._preview_row_index is not None and self._is_plan_sheet_row(self._preview_row_index):
            return self._preview_row_index
        for row_index, _sheet in enumerate(self.parsed_sheets):
            if self._is_plan_sheet_row(row_index):
                return row_index
        return None

    def _is_plan_sheet_row(self, row_index: int) -> bool:
        combo = self.sheet_table.cellWidget(row_index, 5)
        return isinstance(combo, QComboBox) and combo.currentText().strip() == "规划表"

    def _update_execution_mode(self, row_index: int) -> None:
        is_plan = self._is_plan_sheet_row(row_index)
        self.plan_wizard.setVisible(is_plan)
        self.legacy_execution_panel.setVisible(not is_plan)
        self.excel_preview.set_plan_region_visibility(is_plan)

    def _collect_typed_tables(self) -> dict[str, pd.DataFrame]:
        input_tables: dict[str, pd.DataFrame] = {}
        direction = self.direction_combo.currentData()

        for row_index, sheet in enumerate(self.parsed_sheets):
            combo = self.sheet_table.cellWidget(row_index, 5)
            if not isinstance(combo, QComboBox):
                continue
            role = combo.currentText().strip()
            if not role or role == "规划表":
                continue

            sheet_range = self.sheet_ranges.get(row_index, SheetRange())
            start_cell = sheet_range.start_cell or "A1"
            end_cell = sheet_range.end_cell.strip() or _last_cell(sheet.data)
            input_tables[role] = table_from_range(sheet.data, start_cell, end_cell, direction)

        if not input_tables:
            raise ValueError("请至少为一个非规划表 Sheet 指定类型。")
        return input_tables

    def _render_result_preview(self, result: ConversionResult) -> None:
        self.result_tabs.clear()
        for name, table in result.tables.items():
            self.result_tabs.addTab(_dataframe_to_table(table), name)
        self.result_tabs.setVisible(True)

    def export_outputs(self) -> None:
        if not self.current_template or not self.current_result:
            return

        target_dir = QFileDialog.getExistingDirectory(self, "选择导出目录", str(self.output_dir))
        if not target_dir:
            return

        file_names = {output.key: output.file_name for output in self.current_template.outputs}
        file_names["异常清单"] = "异常清单.xlsx"
        try:
            written = write_outputs(target_dir, self.current_result.tables, file_names)
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", str(exc))
            return

        QMessageBox.information(self, "导出完成", "\n".join(str(path) for path in written))

    def _ensure_plan_ranges(self, row_index: int) -> dict[str, SheetRange]:
        if row_index in self.plan_ranges:
            return self.plan_ranges[row_index]

        sheet = self.parsed_sheets[row_index]
        try:
            config = infer_config_from_range(sheet.data, "D2", _last_cell(sheet.data))
            end_row, end_col = coordinate_to_tuple(config.end_cell)
            mechanism = SheetRange(
                config.start_cell,
                f"{get_column_letter(end_col)}{config.activity_row_end}",
            )
            product_start_row = max(1, config.product_row_start - 1)
            product = SheetRange(
                f"{get_column_letter(config.brand_col)}{product_start_row}",
                config.end_cell,
            )
        except Exception:
            mechanism = SheetRange("D2", "")
            product = SheetRange("C10", "")

        self.plan_ranges[row_index] = {"mechanism": mechanism, "product": product}
        return self.plan_ranges[row_index]

    def _display_ranges_for_row(self, row_index: int) -> tuple[str, str]:
        if self.sheet_roles.get(row_index) == "规划表" and row_index in self.plan_ranges:
            ranges = self.plan_ranges[row_index]
            return _format_range(ranges["mechanism"]), _format_range(ranges["product"])
        sheet_range = self.sheet_ranges.get(row_index, SheetRange())
        return sheet_range.start_cell, sheet_range.end_cell or "自动"

    def _update_range_cells(self, row_index: int) -> None:
        if row_index < 0 or row_index >= self.sheet_table.rowCount():
            return
        left, right = self._display_ranges_for_row(row_index)
        left_item = self.sheet_table.item(row_index, 6)
        right_item = self.sheet_table.item(row_index, 7)
        if left_item:
            left_item.setText(left)
            left_item.setToolTip(left)
        if right_item:
            right_item.setText(right)
            right_item.setToolTip(right)

    def _sync_sheet_roles_from_widgets(self) -> None:
        if not hasattr(self, "sheet_table"):
            return
        for row_index in range(self.sheet_table.rowCount()):
            combo = self.sheet_table.cellWidget(row_index, 5)
            if isinstance(combo, QComboBox):
                self.sheet_roles[row_index] = combo.currentText().strip()


def _dataframe_to_table(dataframe: pd.DataFrame) -> QTableWidget:
    max_rows = min(len(dataframe), 500)
    table = QTableWidget(max_rows, len(dataframe.columns))
    table.setHorizontalHeaderLabels([str(column) for column in dataframe.columns])
    table.setAlternatingRowColors(True)
    configure_table_full_text(table)

    for row_index in range(max_rows):
        for column_index, column in enumerate(dataframe.columns):
            value = dataframe.iloc[row_index][column]
            text = "" if pd.isna(value) else str(value)
            item = QTableWidgetItem(text)
            if text:
                item.setToolTip(text)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row_index, column_index, item)

    autosize_table_columns(table)
    return table


def _last_cell(dataframe: pd.DataFrame) -> str:
    if dataframe.empty:
        return "A1"
    return f"{get_column_letter(max(1, len(dataframe.columns)))}{max(1, len(dataframe))}"


def _format_range(sheet_range: SheetRange) -> str:
    if not sheet_range.start_cell:
        return "未设置"
    if not sheet_range.end_cell:
        return f"{sheet_range.start_cell}:自动"
    return f"{sheet_range.start_cell}:{sheet_range.end_cell}"


def run_app(project_root: Path) -> int:
    app = QApplication([])
    window = MainWindow(project_root)
    window.show()
    return app.exec()
