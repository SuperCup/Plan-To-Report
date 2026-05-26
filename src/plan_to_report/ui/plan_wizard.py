from __future__ import annotations

from collections.abc import Callable
from pathlib import Path

import pandas as pd
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFormLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

from ..plan_config import (
    LOGIC_FIELD_OPTIONS,
    ActivityColumnOption,
    AiSettings,
    FieldRowMapping,
    PlanSheetConfig,
    SheetRange,
)
from ..plan_parser import discover_activity_columns, discover_field_rows, infer_config_from_range
from .excel_preview import autosize_table_columns, configure_table_full_text


class PlanWizardWidget(QWidget):
    """规划表配置面板：区域取点、字段映射、活动列勾选与商品区参数。"""

    pick_requested = Signal(str, str)  # region_key, edge

    def __init__(
        self,
        parent: QWidget | None = None,
        project_root: Path | None = None,
        on_save_ai: Callable[[AiSettings], None] | None = None,
    ) -> None:
        super().__init__(parent)
        self._project_root = project_root
        self._on_save_ai = on_save_ai
        self._config = PlanSheetConfig()
        self._raw: pd.DataFrame | None = None
        self._mechanism_range = SheetRange("D2", "")
        self._product_range = SheetRange("C10", "")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("规划表配置")
        title.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(title)

        layout.addWidget(self._build_range_group())
        layout.addWidget(self._build_mapping_tabs(), stretch=1)
        layout.addWidget(self._build_product_group())
        layout.addWidget(self._build_ai_group())

        self._refresh_range_labels()

    def _build_range_group(self) -> QWidget:
        group = QGroupBox("区域划分")
        layout = QVBoxLayout(group)

        self.mechanism_range_label = QLabel()
        self.product_range_label = QLabel()
        layout.addWidget(self.mechanism_range_label)
        layout.addLayout(
            self._pick_button_row(
                "机制信息区",
                lambda: self.pick_requested.emit("mechanism", "start"),
                lambda: self.pick_requested.emit("mechanism", "end"),
            )
        )
        layout.addWidget(self.product_range_label)
        layout.addLayout(
            self._pick_button_row(
                "商品勾选区",
                lambda: self.pick_requested.emit("product", "start"),
                lambda: self.pick_requested.emit("product", "end"),
            )
        )

        self.range_summary = QLabel("请选择规划表 Sheet 后，在中央预览区框选两块区域。")
        self.range_summary.setWordWrap(True)
        self.range_summary.setStyleSheet("color: #666;")
        layout.addWidget(self.range_summary)
        return group

    def _pick_button_row(self, label: str, start_callback: Callable[[], None], end_callback: Callable[[], None]) -> QHBoxLayout:
        row = QHBoxLayout()
        row.addWidget(QLabel(label))
        start_btn = QPushButton("左上")
        end_btn = QPushButton("右下")
        start_btn.clicked.connect(start_callback)
        end_btn.clicked.connect(end_callback)
        row.addWidget(start_btn)
        row.addWidget(end_btn)
        row.addStretch()
        return row

    def _build_mapping_tabs(self) -> QWidget:
        tabs = QTabWidget()

        field_page = QWidget()
        field_layout = QVBoxLayout(field_page)
        self.field_table = QTableWidget(0, 3)
        self.field_table.setHorizontalHeaderLabels(["行", "字段文本", "映射"])
        configure_table_full_text(self.field_table)
        field_layout.addWidget(self.field_table)
        refresh_fields = QPushButton("按当前机制区重新识别字段")
        refresh_fields.clicked.connect(self._rebuild_from_ui)
        field_layout.addWidget(refresh_fields)

        column_page = QWidget()
        column_layout = QVBoxLayout(column_page)
        bar = QHBoxLayout()
        select_all = QPushButton("全选")
        select_none = QPushButton("全不选")
        select_all.clicked.connect(lambda: self._set_all_columns(True))
        select_none.clicked.connect(lambda: self._set_all_columns(False))
        bar.addWidget(select_all)
        bar.addWidget(select_none)
        bar.addStretch()
        column_layout.addLayout(bar)
        self.column_table = QTableWidget(0, 3)
        self.column_table.setHorizontalHeaderLabels(["生成", "列", "摘要"])
        configure_table_full_text(self.column_table)
        column_layout.addWidget(self.column_table)

        tabs.addTab(field_page, "字段映射")
        tabs.addTab(column_page, "活动列")
        return tabs

    def _build_product_group(self) -> QWidget:
        group = QGroupBox("商品矩阵参数")
        form = QFormLayout(group)
        self.product_row_spin = QSpinBox()
        self.product_row_spin.setRange(1, 5000)
        self.product_row_spin.setValue(11)
        self.brand_col_spin = QSpinBox()
        self.brand_col_spin.setRange(1, 100)
        self.brand_col_spin.setValue(3)
        self.spec_col_spin = QSpinBox()
        self.spec_col_spin.setRange(1, 100)
        self.spec_col_spin.setValue(4)
        form.addRow("商品起始行：", self.product_row_spin)
        form.addRow("子品牌列：", self.brand_col_spin)
        form.addRow("规格列：", self.spec_col_spin)
        return group

    def _build_ai_group(self) -> QWidget:
        group = QGroupBox("DeepSeek 辅助")
        form = QFormLayout(group)
        self.ai_enable_check = QCheckBox("启用机制拆分 / 歧义口味解析")
        self.ai_key_edit = QLineEdit()
        self.ai_key_edit.setPlaceholderText("API Key")
        self.ai_key_edit.setEchoMode(QLineEdit.Password)
        self.ai_model_edit = QLineEdit("deepseek-chat")
        self.ai_base_url_edit = QLineEdit("https://api.deepseek.com")
        self.ai_config_hint = QLabel("")
        self.ai_config_hint.setWordWrap(True)
        self.ai_config_hint.setStyleSheet("color: #666;")
        self.save_ai_btn = QPushButton("保存配置")
        self.save_ai_btn.clicked.connect(self._save_ai_settings_clicked)
        self.show_key_btn = QPushButton("显示")
        self.show_key_btn.setCheckable(True)
        self.show_key_btn.toggled.connect(self._toggle_key_visibility)

        key_row = QHBoxLayout()
        key_row.addWidget(self.ai_key_edit, stretch=1)
        key_row.addWidget(self.show_key_btn)
        form.addRow(self.ai_enable_check)
        form.addRow("Key：", key_row)
        form.addRow("模型：", self.ai_model_edit)
        form.addRow("地址：", self.ai_base_url_edit)
        form.addRow(self.save_ai_btn)
        form.addRow(self.ai_config_hint)
        return group

    def refresh_from_sheet(
        self,
        raw: pd.DataFrame,
        mechanism_range: SheetRange | str,
        product_range: SheetRange | None = None,
    ) -> None:
        self._raw = raw
        if isinstance(mechanism_range, SheetRange):
            self._mechanism_range = mechanism_range
            self._product_range = product_range or self._product_range
        else:
            self._mechanism_range = SheetRange(mechanism_range or "D2", "")
            if isinstance(product_range, SheetRange):
                self._product_range = product_range

        if raw is None or raw.empty:
            self.range_summary.setText("当前 Sheet 数据为空。")
            return

        self._rebuild_config(preserve_ui=True)
        self._refresh_range_labels()
        self._refresh_field_table()
        self._refresh_column_table()

    def get_config(self) -> PlanSheetConfig:
        self._sync_config_from_ui()
        return self._config

    def is_ready(self) -> tuple[bool, str]:
        config = self.get_config()
        if not self._mechanism_range.start_cell:
            return False, "请先框选机制信息区域。"
        if not self._product_range.start_cell:
            return False, "请先框选商品勾选区域。"
        if not config.selected_column_indices():
            return False, "请至少勾选一个活动列。"
        if not [m for m in config.field_mappings if m.field_key]:
            return False, "请至少映射一个活动字段行。"
        return True, ""

    def load_ai_settings(self, settings: AiSettings, config_path: Path | None = None) -> None:
        self.ai_enable_check.setChecked(settings.enabled)
        self.ai_key_edit.setText(settings.api_key)
        self.ai_model_edit.setText(settings.model or "deepseek-chat")
        self.ai_base_url_edit.setText(settings.base_url or "https://api.deepseek.com")
        self._config.ai_settings = settings
        if config_path and config_path.exists():
            self.ai_config_hint.setText(f"配置文件：{config_path}")
        else:
            self.ai_config_hint.setText("未找到配置文件，可填写后保存；Key 也可来自 DEEPSEEK_API_KEY。")

    def get_ai_settings(self) -> AiSettings:
        return AiSettings(
            enabled=self.ai_enable_check.isChecked(),
            api_key=self.ai_key_edit.text().strip(),
            model=self.ai_model_edit.text().strip() or "deepseek-chat",
            base_url=self.ai_base_url_edit.text().strip() or "https://api.deepseek.com",
        )

    def _rebuild_from_ui(self) -> None:
        self._rebuild_config(preserve_ui=True)
        self._refresh_range_labels()
        self._refresh_field_table()
        self._refresh_column_table()

    def _rebuild_config(self, preserve_ui: bool) -> None:
        if self._raw is None or self._raw.empty:
            return

        previous_mappings = {mapping.excel_row: mapping.field_key for mapping in self._read_field_mappings_from_table()}
        previous_columns = self._read_column_selection_from_table()
        self._config = _config_from_ranges(self._raw, self._mechanism_range, self._product_range)

        if preserve_ui:
            for mapping in self._config.field_mappings:
                if mapping.excel_row in previous_mappings:
                    mapping.field_key = previous_mappings[mapping.excel_row]
            for column in self._config.activity_columns:
                if column.excel_col in previous_columns:
                    column.selected = previous_columns[column.excel_col]

        self.product_row_spin.setValue(self._config.product_row_start)
        self.brand_col_spin.setValue(self._config.brand_col)
        self.spec_col_spin.setValue(self._config.spec_col)
        self._config.ai_settings = self.get_ai_settings()

    def _sync_config_from_ui(self) -> None:
        self._config.product_row_start = self.product_row_spin.value()
        self._config.brand_col = self.brand_col_spin.value()
        self._config.spec_col = self.spec_col_spin.value()

        mappings = self._read_field_mappings_from_table()
        if mappings:
            self._config.field_mappings = mappings

        columns: list[ActivityColumnOption] = []
        for row in range(self.column_table.rowCount()):
            check = self.column_table.cellWidget(row, 0)
            col_item = self.column_table.item(row, 1)
            preview_item = self.column_table.item(row, 2)
            if not isinstance(check, QCheckBox) or not col_item or not preview_item:
                continue
            letter = col_item.text()
            columns.append(
                ActivityColumnOption(
                    excel_col=_column_letter_to_index(letter),
                    column_letter=letter,
                    preview=preview_item.text(),
                    selected=check.isChecked(),
                )
            )
        if columns:
            self._config.activity_columns = columns

        self._config.ai_settings = self.get_ai_settings()

    def _refresh_range_labels(self) -> None:
        self.mechanism_range_label.setText(f"机制信息区：{_format_range(self._mechanism_range)}")
        self.product_range_label.setText(f"商品勾选区：{_format_range(self._product_range)}")
        if self._raw is not None and not self._raw.empty:
            self.range_summary.setText(
                f"活动字段行 {self._config.activity_row_start}-{self._config.activity_row_end}；"
                f"商品从第 {self._config.product_row_start} 行开始；"
                f"活动列 {len(self._config.activity_columns)} 个。"
            )

    def _refresh_field_table(self) -> None:
        mappings = self._config.field_mappings
        self.field_table.setRowCount(len(mappings))
        for row_index, mapping in enumerate(mappings):
            row_item = QTableWidgetItem(str(mapping.excel_row))
            row_item.setFlags(row_item.flags() & ~Qt.ItemIsEditable)
            label_item = QTableWidgetItem(mapping.label_text)
            label_item.setFlags(label_item.flags() & ~Qt.ItemIsEditable)
            self.field_table.setItem(row_index, 0, row_item)
            self.field_table.setItem(row_index, 1, label_item)

            combo = QComboBox()
            combo.addItems(LOGIC_FIELD_OPTIONS)
            if mapping.field_key in LOGIC_FIELD_OPTIONS:
                combo.setCurrentText(mapping.field_key)
            self.field_table.setCellWidget(row_index, 2, combo)
        autosize_table_columns(self.field_table, max_width=260)

    def _refresh_column_table(self) -> None:
        columns = self._config.activity_columns
        self.column_table.setRowCount(len(columns))
        for row_index, column in enumerate(columns):
            check = QCheckBox()
            check.setChecked(column.selected)
            self.column_table.setCellWidget(row_index, 0, check)
            col_item = QTableWidgetItem(column.column_letter)
            col_item.setFlags(col_item.flags() & ~Qt.ItemIsEditable)
            preview_item = QTableWidgetItem(column.preview)
            preview_item.setFlags(preview_item.flags() & ~Qt.ItemIsEditable)
            self.column_table.setItem(row_index, 1, col_item)
            self.column_table.setItem(row_index, 2, preview_item)
        autosize_table_columns(self.column_table, max_width=320)

    def _read_field_mappings_from_table(self) -> list[FieldRowMapping]:
        mappings: list[FieldRowMapping] = []
        for row in range(self.field_table.rowCount()):
            row_item = self.field_table.item(row, 0)
            label_item = self.field_table.item(row, 1)
            combo = self.field_table.cellWidget(row, 2)
            if not row_item or not label_item or not isinstance(combo, QComboBox):
                continue
            mappings.append(
                FieldRowMapping(
                    excel_row=int(row_item.text()),
                    label_text=label_item.text(),
                    field_key=combo.currentText().strip(),
                )
            )
        return mappings

    def _read_column_selection_from_table(self) -> dict[int, bool]:
        selected: dict[int, bool] = {}
        for row in range(self.column_table.rowCount()):
            check = self.column_table.cellWidget(row, 0)
            col_item = self.column_table.item(row, 1)
            if not isinstance(check, QCheckBox) or not col_item:
                continue
            selected[_column_letter_to_index(col_item.text())] = check.isChecked()
        return selected

    def _set_all_columns(self, selected: bool) -> None:
        for row in range(self.column_table.rowCount()):
            check = self.column_table.cellWidget(row, 0)
            if isinstance(check, QCheckBox):
                check.setChecked(selected)

    def _save_ai_settings_clicked(self) -> None:
        if self._on_save_ai:
            self._on_save_ai(self.get_ai_settings())

    def _toggle_key_visibility(self, visible: bool) -> None:
        self.ai_key_edit.setEchoMode(QLineEdit.Normal if visible else QLineEdit.Password)
        self.show_key_btn.setText("隐藏" if visible else "显示")


def _config_from_ranges(raw: pd.DataFrame, mechanism_range: SheetRange, product_range: SheetRange) -> PlanSheetConfig:
    mechanism_start = mechanism_range.start_cell or "D2"
    product_start = product_range.start_cell or ""
    inferred_end = product_range.end_cell or _last_cell(raw)

    try:
        base_config = infer_config_from_range(raw, mechanism_start, inferred_end)
    except Exception:
        base_config = PlanSheetConfig()

    mech_start_row, mech_start_col = _parse_or_default(mechanism_start, (base_config.activity_row_start, base_config.label_col))
    if mechanism_range.end_cell:
        mech_end_row, mech_end_col = _parse_or_default(
            mechanism_range.end_cell,
            (base_config.activity_row_end, _column_from_cell(base_config.end_cell)),
        )
    else:
        mech_end_row = base_config.activity_row_end
        mech_end_col = _column_from_cell(base_config.end_cell)

    if product_start:
        product_start_row, product_start_col = _parse_or_default(product_start, (base_config.product_row_start, base_config.brand_col))
        brand_col = product_start_col
        spec_col = product_start_col + 1
        product_row_start = product_start_row
        if product_start_col >= mech_start_col + 1:
            brand_col = base_config.brand_col
            spec_col = base_config.spec_col
        if _looks_like_product_header(raw, product_start_row, brand_col, spec_col):
            product_row_start += 1
    else:
        product_row_start = base_config.product_row_start
        brand_col = base_config.brand_col
        spec_col = base_config.spec_col

    product_end_row, product_end_col = _parse_or_default(
        product_range.end_cell or base_config.end_cell or _last_cell(raw),
        (len(raw), len(raw.columns)),
    )
    end_row = max(product_end_row, mech_end_row)
    end_col = max(product_end_col, mech_end_col)

    config = PlanSheetConfig(
        start_cell=mechanism_start,
        end_cell=f"{get_column_letter(end_col)}{end_row}",
        label_col=mech_start_col,
        first_activity_col=mech_start_col + 1,
        activity_row_start=mech_start_row,
        activity_row_end=mech_end_row,
        product_row_start=product_row_start,
        brand_col=brand_col,
        spec_col=spec_col,
    )
    config.field_mappings = discover_field_rows(raw, config)
    config.activity_columns = discover_activity_columns(raw, config)
    return config


def _looks_like_product_header(raw: pd.DataFrame, row: int, brand_col: int, spec_col: int) -> bool:
    brand_text = _cell_text(raw, row, brand_col)
    spec_text = _cell_text(raw, row, spec_col)
    return brand_text in {"品牌", "子品牌"} or "规格" in spec_text


def _cell_text(raw: pd.DataFrame, row: int, col: int) -> str:
    if row < 1 or col < 1 or row - 1 >= len(raw) or col - 1 >= len(raw.columns):
        return ""
    value = raw.iat[row - 1, col - 1]
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    return str(value).strip()


def _parse_or_default(cell: str, default: tuple[int, int]) -> tuple[int, int]:
    try:
        return coordinate_to_tuple(cell.upper())
    except ValueError:
        return default


def _column_from_cell(cell: str) -> int:
    if not cell:
        return 1
    return _parse_or_default(cell, (1, 1))[1]


def _column_letter_to_index(letter: str) -> int:
    from openpyxl.utils.cell import column_index_from_string

    return column_index_from_string(letter)


def _last_cell(dataframe: pd.DataFrame) -> str:
    if dataframe.empty:
        return "A1"
    return f"{get_column_letter(max(1, len(dataframe.columns)))}{max(1, len(dataframe))}"


def _format_range(sheet_range: SheetRange) -> str:
    if not sheet_range.start_cell:
        return "未设置"
    if not sheet_range.end_cell:
        return f"{sheet_range.start_cell}:自动"
    return f"{sheet_range.start_cell}:{sheet_range.end_cell}"
