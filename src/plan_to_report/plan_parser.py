from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from .deepseek_assist import (
    DeepseekAssistError,
    assist_split_mechanism,
    should_try_ai_mechanism_split,
)
from .models import ConversionIssue
from .plan_config import REMARK_FIELD_ORDER, AiSettings, FieldRowMapping, PlanSheetConfig


@dataclass
class ParsedActivity:
    column_letter: str
    column_index: int
    fields: dict[str, Any]
    remark: str
    split_index: int = 0
    split_total: int = 1


@dataclass
class PlanParseResult:
    activities: list[ParsedActivity] = field(default_factory=list)
    issues: list[ConversionIssue] = field(default_factory=list)


_LABEL_ALIASES: dict[str, list[str]] = {
    "活动类型": ["活动类型"],
    "活动机制": ["活动机制"],
    "机制力度": ["机制力度", "力度"],
    "活动时间": ["活动时间", "时间"],
    "活动区域/渠道": ["活动区域/渠道", "活动区域", "渠道"],
    "优惠券类型": ["优惠券类型", "券类型"],
    "预算分配": ["预算分配", "预算分配(元)", "预算"],
    "机制说明": ["机制说明", "说明"],
}


def infer_config_from_range(
    raw: pd.DataFrame,
    start_cell: str,
    end_cell: str,
) -> PlanSheetConfig:
    start_row, start_col = coordinate_to_tuple(start_cell.upper())
    end_row, end_col = _resolve_end(raw, start_cell, end_cell)
    product_row = _detect_product_row_start(raw, start_row, end_row, start_col)
    activity_row_end = max(start_row, product_row - 1)

    config = PlanSheetConfig(
        start_cell=start_cell.upper(),
        end_cell=f"{get_column_letter(end_col)}{end_row}",
        label_col=start_col,
        first_activity_col=start_col + 1,
        activity_row_start=start_row,
        activity_row_end=activity_row_end,
        product_row_start=product_row,
        brand_col=3,
        spec_col=4,
    )
    config.field_mappings = discover_field_rows(raw, config)
    config.activity_columns = discover_activity_columns(raw, config)
    return config


def discover_field_rows(raw: pd.DataFrame, config: PlanSheetConfig) -> list[FieldRowMapping]:
    mappings: list[FieldRowMapping] = []
    for row in range(config.activity_row_start, config.activity_row_end + 1):
        label = _cell_text(raw, row, config.label_col)
        if label:
            mappings.append(FieldRowMapping(excel_row=row, label_text=label, field_key=_guess_field_key(label)))
        else:
            mappings.append(
                FieldRowMapping(
                    excel_row=row,
                    label_text="(空白行，可映射为机制说明)",
                    field_key="机制说明",
                )
            )
    return mappings


def discover_activity_columns(raw: pd.DataFrame, config: PlanSheetConfig) -> list:
    from .plan_config import ActivityColumnOption

    _, end_col = coordinate_to_tuple(config.end_cell)
    options: list[ActivityColumnOption] = []
    mechanism_row = next((m.excel_row for m in config.field_mappings if m.field_key == "活动机制"), None)
    type_row = next((m.excel_row for m in config.field_mappings if m.field_key == "活动类型"), None)

    for col in range(config.first_activity_col, end_col + 1):
        preview_parts: list[str] = []
        if type_row:
            value = _cell_text(raw, type_row, col)
            if value:
                preview_parts.append(value)
        if mechanism_row:
            value = _cell_text(raw, mechanism_row, col)
            if value:
                preview_parts.append(value)
        preview = " / ".join(preview_parts) if preview_parts else "(空列)"
        has_content = preview != "(空列)"
        options.append(
            ActivityColumnOption(
                excel_col=col,
                column_letter=get_column_letter(col),
                preview=preview,
                selected=has_content,
            )
        )
    return options


def parse_plan_sheet(raw: pd.DataFrame, config: PlanSheetConfig) -> PlanParseResult:
    result = PlanParseResult()
    selected_cols = config.selected_column_indices()
    if not selected_cols:
        result.issues.append(
            ConversionIssue("error", "规划表", None, None, "请至少勾选一个需要生成的活动列。")
        )
        return result

    mapped_rows = [m for m in config.field_mappings if m.field_key]
    if not mapped_rows:
        result.issues.append(
            ConversionIssue("error", "规划表", None, None, "请至少映射一个活动参数字段行。")
        )
        return result

    for col in selected_cols:
        fields: dict[str, Any] = {"_column": get_column_letter(col)}
        for mapping in mapped_rows:
            fields[mapping.field_key] = _cell_value(raw, mapping.excel_row, col)

        if _is_blank(fields.get("活动机制")) and _is_blank(fields.get("活动类型")):
            result.issues.append(
                ConversionIssue(
                    "info",
                    "规划表",
                    None,
                    get_column_letter(col),
                    f"列 {get_column_letter(col)} 活动类型与活动机制均为空，已跳过。",
                )
            )
            continue

        mechanism = fields.get("活动机制")
        segments = (
            split_mechanisms(mechanism, config.ai_settings, get_column_letter(col), result.issues)
            if not _is_blank(mechanism)
            else [mechanism]
        )
        segments = segments or [mechanism]
        if len(segments) > 1:
            result.issues.append(
                ConversionIssue(
                    "info",
                    "活动汇总表",
                    None,
                    get_column_letter(col),
                    f"列 {get_column_letter(col)} 活动机制拆分为 {len(segments)} 条活动。",
                )
            )

        for index, segment in enumerate(segments):
            activity_fields = dict(fields)
            if segment is not None and not _is_blank(segment):
                activity_fields["活动机制"] = segment
            remark = build_remark(activity_fields)
            result.activities.append(
                ParsedActivity(
                    column_letter=get_column_letter(col),
                    column_index=col,
                    fields=activity_fields,
                    remark=remark,
                    split_index=index,
                    split_total=len(segments),
                )
            )

    return result


def split_mechanisms(
    value: Any,
    ai_settings: AiSettings | None = None,
    column_letter: str | None = None,
    issues: list[ConversionIssue] | None = None,
) -> list[str]:
    if _is_blank(value):
        return []
    text = str(value).strip()
    rule_segments = _split_mechanisms_by_rules(text)
    if len(rule_segments) > 1:
        return rule_segments

    settings = ai_settings or AiSettings()
    if settings.enabled and should_try_ai_mechanism_split(text, rule_segments):
        try:
            ai_segments = assist_split_mechanism(text, settings)
            if ai_segments and len(ai_segments) > 1:
                if issues is not None:
                    issues.append(
                        ConversionIssue(
                            "info",
                            "活动汇总表",
                            None,
                            column_letter,
                            f"列 {column_letter} 活动机制由 AI 拆分为 {len(ai_segments)} 段。",
                        )
                    )
                return ai_segments
        except DeepseekAssistError as exc:
            if issues is not None:
                issues.append(
                    ConversionIssue("warning", "活动汇总表", None, column_letter, str(exc))
                )

    return rule_segments


def _split_mechanisms_by_rules(text: str) -> list[str]:
    if re.search(r"[①②③④⑤⑥]", text):
        parts = re.split(r"[①②③④⑤⑥]\s*", text)
        cleaned = [part.strip() for part in parts if part and part.strip()]
        if cleaned:
            return cleaned
    matches = re.findall(r"满\s*\d+\s*[-减]\s*\d+", text)
    if len(matches) > 1:
        return [match.strip() for match in matches]
    return [text]


def build_remark(fields: dict[str, Any]) -> str:
    parts: list[str] = []
    for key in REMARK_FIELD_ORDER:
        value = fields.get(key)
        if not _is_blank(value):
            parts.append(str(value).strip())
    return "+".join(parts)


def _guess_field_key(label: str) -> str:
    normalized = label.strip()
    for field_key, aliases in _LABEL_ALIASES.items():
        for alias in aliases:
            if alias in normalized or normalized in alias:
                return field_key
    return ""


def _detect_product_row_start(
    raw: pd.DataFrame,
    start_row: int,
    end_row: int,
    label_col: int,
) -> int:
    for row in range(start_row + 1, end_row + 1):
        label = _cell_text(raw, row, label_col)
        if label and "规格" in label:
            return row + 1
    return min(start_row + 9, end_row)


def _resolve_end(raw: pd.DataFrame, start_cell: str, end_cell: str) -> tuple[int, int]:
    start_row, start_col = coordinate_to_tuple(start_cell.upper())
    if end_cell and end_cell.strip():
        end_row, end_col = coordinate_to_tuple(end_cell.upper())
        return end_row, end_col
    return len(raw), len(raw.columns)


def _cell_value(raw: pd.DataFrame, row: int, col: int) -> Any:
    if row < 1 or col < 1:
        return None
    row_index = row - 1
    col_index = col - 1
    if row_index >= len(raw) or col_index >= len(raw.columns):
        return None
    value = raw.iat[row_index, col_index]
    if _is_blank(value):
        return None
    return value


def _cell_text(raw: pd.DataFrame, row: int, col: int) -> str:
    value = _cell_value(raw, row, col)
    if value is None:
        return ""
    return str(value).strip()


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except TypeError:
        return False
