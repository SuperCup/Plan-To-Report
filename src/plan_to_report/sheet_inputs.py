from __future__ import annotations

import pandas as pd
from openpyxl.utils.cell import get_column_letter
from PySide6.QtWidgets import QComboBox, QTableWidget

from .excel_io import ParsedSheet, table_from_range
from .plan_config import SheetRange


def load_role_table(
    parsed_sheets: list[ParsedSheet],
    sheet_table: QTableWidget,
    sheet_ranges: dict[int, SheetRange],
    role: str,
    direction: str = "row",
) -> pd.DataFrame | None:
    for row_index, sheet in enumerate(parsed_sheets):
        combo = sheet_table.cellWidget(row_index, 5)
        if not isinstance(combo, QComboBox):
            continue
        if combo.currentText().strip() != role:
            continue

        sheet_range = sheet_ranges.get(row_index, SheetRange())
        if _is_default_range(sheet_range):
            sheet_range = _detect_role_range(sheet.data, role)
        start_cell = sheet_range.start_cell or "A1"
        end_cell = sheet_range.end_cell.strip() or _last_cell(sheet.data)
        return table_from_range(sheet.data, start_cell, end_cell, direction)
    return None


def _last_cell(dataframe: pd.DataFrame) -> str:
    if dataframe.empty:
        return "A1"
    return f"{get_column_letter(max(1, len(dataframe.columns)))}{max(1, len(dataframe))}"


def _is_default_range(sheet_range: SheetRange) -> bool:
    return (sheet_range.start_cell or "A1").upper() == "A1" and not sheet_range.end_cell.strip()


def _detect_role_range(dataframe: pd.DataFrame, role: str) -> SheetRange:
    if dataframe.empty:
        return SheetRange()
    if role == "商品清单":
        header_row = _find_header_row(
            dataframe,
            required_groups=[
                ("商品条形码", "UPC条形码", "条形码"),
                ("标品名称",),
                ("品牌名称",),
                ("规格名称",),
                ("口味名称", "口味"),
            ],
        )
        if header_row is not None:
            return _range_from_header_row(dataframe, header_row)
    if role == "商品匹配逻辑":
        header_row = _find_header_row(
            dataframe,
            required_groups=[
                ("品牌",),
                ("规格",),
                ("选品逻辑", "选品逻辑说明", "逻辑"),
            ],
        )
        if header_row is not None:
            return _range_from_header_row(dataframe, header_row)
    return SheetRange("A1", "")


def _find_header_row(dataframe: pd.DataFrame, required_groups: list[tuple[str, ...]]) -> int | None:
    max_scan_rows = min(len(dataframe), 30)
    for row_index in range(max_scan_rows):
        values = [_normalize_text(value) for value in dataframe.iloc[row_index].tolist()]
        if all(any(any(candidate in value for value in values) for candidate in group) for group in required_groups):
            return row_index + 1
    return None


def _range_from_header_row(dataframe: pd.DataFrame, header_row: int) -> SheetRange:
    row_values = dataframe.iloc[header_row - 1].tolist()
    non_blank_columns = [
        index + 1
        for index, value in enumerate(row_values)
        if _normalize_text(value)
    ]
    start_col = min(non_blank_columns) if non_blank_columns else 1
    end_col = _last_non_empty_column(dataframe, header_row, start_col)
    end_row = _last_non_empty_row(dataframe, header_row, start_col, end_col)
    return SheetRange(f"{get_column_letter(start_col)}{header_row}", f"{get_column_letter(end_col)}{end_row}")


def _last_non_empty_column(dataframe: pd.DataFrame, start_row: int, start_col: int) -> int:
    last_col = start_col
    for col_index in range(start_col - 1, len(dataframe.columns)):
        if dataframe.iloc[start_row - 1 :, col_index].map(_normalize_text).astype(bool).any():
            last_col = col_index + 1
    return last_col


def _last_non_empty_row(dataframe: pd.DataFrame, start_row: int, start_col: int, end_col: int) -> int:
    last_row = start_row
    for row_index in range(start_row - 1, len(dataframe)):
        values = dataframe.iloc[row_index, start_col - 1 : end_col].map(_normalize_text)
        if values.astype(bool).any():
            last_row = row_index + 1
    return last_row


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()
