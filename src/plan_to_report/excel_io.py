from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
import pandas as pd


WorkbookData = dict[str, pd.DataFrame]


@dataclass
class ParsedSheet:
    file_path: Path
    sheet_name: str
    data: pd.DataFrame
    merge_notes: list[str] = field(default_factory=list)

    @property
    def source_id(self) -> str:
        return f"{self.file_path.name}::{self.sheet_name}"


def read_workbook(path: str | Path) -> WorkbookData:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"文件不存在：{workbook_path}")
    return pd.read_excel(workbook_path, sheet_name=None, dtype=object)


def parse_workbook_sheets(path: str | Path) -> list[ParsedSheet]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"文件不存在：{workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    parsed: list[ParsedSheet] = []
    for worksheet in workbook.worksheets:
        values = [[cell.value for cell in row] for row in worksheet.iter_rows()]
        notes = _split_merged_cells(values, worksheet.merged_cells.ranges)
        parsed.append(
            ParsedSheet(
                file_path=workbook_path,
                sheet_name=worksheet.title,
                data=pd.DataFrame(values, dtype=object),
                merge_notes=notes,
            )
        )
    return parsed


def table_from_range(raw_table: pd.DataFrame, start_cell: str, end_cell: str, direction: str) -> pd.DataFrame:
    start_row, start_column = coordinate_to_tuple(start_cell.upper())
    end_row, end_column = coordinate_to_tuple(end_cell.upper())
    if start_row > end_row or start_column > end_column:
        raise ValueError("结束位置必须在起始位置的右下方。")

    sliced = raw_table.iloc[start_row - 1 : end_row, start_column - 1 : end_column].copy()
    sliced = sliced.reset_index(drop=True)
    sliced.columns = range(len(sliced.columns))

    if direction == "column":
        sliced = sliced.transpose().reset_index(drop=True)

    if sliced.empty:
        return pd.DataFrame()

    header = _dedupe_headers([_normalize_header(value, index) for index, value in enumerate(sliced.iloc[0].tolist())])
    records = sliced.iloc[1:].reset_index(drop=True)
    records.columns = header
    return records.dropna(how="all").reset_index(drop=True)


def first_sheet(workbook: WorkbookData) -> tuple[str, pd.DataFrame]:
    if not workbook:
        raise ValueError("Excel 文件中没有可读取的 sheet。")
    name = next(iter(workbook))
    return name, workbook[name]


def get_sheet(workbook: WorkbookData, sheet_name: str | None) -> tuple[str, pd.DataFrame]:
    if sheet_name and sheet_name in workbook:
        return sheet_name, workbook[sheet_name]
    if sheet_name:
        raise KeyError(f"找不到 sheet：{sheet_name}")
    return first_sheet(workbook)


def write_outputs(output_dir: str | Path, tables: dict[str, pd.DataFrame], file_names: dict[str, str]) -> list[Path]:
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)

    written: list[Path] = []
    for key, table in tables.items():
        file_name = file_names.get(key, f"{key}.xlsx")
        target_path = target_dir / file_name
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            table.to_excel(writer, index=False, sheet_name=key[:31])
        written.append(target_path)
    return written


def _split_merged_cells(values: list[list[Any]], merged_ranges: Any) -> list[str]:
    notes: list[str] = []
    for merged_range in list(merged_ranges):
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        original = _get_matrix_value(values, min_row, min_col)
        row_count = max_row - min_row + 1
        column_count = max_col - min_col + 1

        if row_count == 1 and column_count > 1:
            for column in range(min_col, max_col + 1):
                _set_matrix_value(values, min_row, column, original)
            continue

        if column_count == 1 and row_count > 1:
            for row in range(min_row, max_row + 1):
                _set_matrix_value(values, row, min_col, None)
            _set_matrix_value(values, max_row, min_col, original)
            continue

        if row_count > 1 and column_count > 1:
            notes.append(f"{merged_range.coord} 为多行多列合并单元格，需要用户补充拆分后的内容。")

    return notes


def _get_matrix_value(values: list[list[Any]], row: int, column: int) -> Any:
    if row - 1 >= len(values) or column - 1 >= len(values[row - 1]):
        return None
    return values[row - 1][column - 1]


def _set_matrix_value(values: list[list[Any]], row: int, column: int, value: Any) -> None:
    while len(values) < row:
        values.append([])
    while len(values[row - 1]) < column:
        values[row - 1].append(None)
    values[row - 1][column - 1] = value


def _normalize_header(value: Any, index: int) -> str:
    if value is None or pd.isna(value):
        return f"未命名字段{index + 1}"
    return str(value).strip() or f"未命名字段{index + 1}"


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        if count == 0:
            unique.append(header)
        else:
            unique.append(f"{header}_{count + 1}")
    return unique
